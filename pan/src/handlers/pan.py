import re
import os
import json
import datetime

from utils import req
from config import settings
from openpyxl import load_workbook

class PanHandler(object):

    #  初始化并定义客户端socket连接
    def __init__(self,conn):
        self.conn = conn
        self.username = None

    # 路径
    def home_path(self):
        return os.path.join(settings.USER_FOLDER_PATH, self.username)

    #发送数据
    def send_json_data(self,**kwargs):
        req.send_data(self.conn,json.dumps(kwargs))

    def send_file_by_seek(self,file_size,file_path,seek=0):
        req.send_file_by_seek(self.conn,file_size,file_path,seek)

    def recv_save_file(self, target_file_path):
        req.recv_save_file(self.conn, target_file_path)

    def send_file_by_seek(self, file_size, file_path, seek=0):
        req.send_file_by_seek(self.conn, file_size, file_path, seek)

    # 读取xlsx文件判断用户能否登录
    def login(self,username,password):
        wb = load_workbook(settings.XLSX_FILE_PATH)
        sheet = wb.worksheets[0]

        success = False   # 账号密码正确的开关
        for row in sheet.iter_rows(2):
            if username == row[0].value and password == row[1].value:
                success = True
                break
        if success:
            self.send_json_data(status=True,error="登陆成功")
            self.username = username
        else:
            self.send_json_data(status=False,error="登陆失败")

    # 用户注册检测不存在的数据写入xlsx文件里面
    def register(self,username,password):

        wb = load_workbook(settings.XLSX_FILE_PATH)
        sheet = wb.worksheets[0]

        # 监测用户是否已经存在
        exists =False
        for row in sheet.iter_rows(2):
            if username == row[0].value:
                exists = True
                break
        if exists:
            self.send_json_data(status=False,error='用户名已存在')
            return

        # 注册用户并写入xlsx
        max_row = sheet.max_row
        data_list = [username,password,datetime.datetime.now().strftime("%Y-%m-%d")]
        for i, item in enumerate(data_list,1):
            cell = sheet.cell(max_row+1,i)
            cell.value = item
        wb.save(settings.XLSX_FILE_PATH)

        # 创建用户目录
        user_folder = os.path.join(settings.USER_FOLDER_PATH, username)
        os.makedirs(user_folder)

        self.send_json_data(status=True, data='注册成功')

    def ls(self,folder_path=None):
        """ 查看当前用户目录下所有的文件
                    1. folder_path=None，查看用户根目录
                    2. folder_path不为空，查看  用户目录/folder_path 中的文件
        """
        if not self.username:
            self.send_json_data(status=False,error="登录后才能查看")
            return
        if not folder_path:
            # 根目录: files目录+用户名
            data = "\n".join(os.listdir(self.home_path()))
            self.send_json_data(status=True,data=data)
            return

        target_folder = os.path.join(self.home_path(),folder_path)
        if not os.path.exists(target_folder):
            self.send_json_data(status=False,error="路径不存在")
            return
        if not os.path.isdir(target_folder):
            self.send_json_data(status=False,error="文件夹不存在")
            return

        data = "\n".join(os.listdir(target_folder))
        self.send_json_data(status=True,data=data)

     #  上传文件
    def upload(self,file_path):

        # 用户未登录
        if not self.username:
            self.send_json_data(status=False, error="登录后才能查看")
            return

        target_file_path = os.path.join(self.home_path(),file_path)
        folder = os.path.dirname(target_file_path)
        if not os.path.exists(folder):
            os.makedirs(folder)

        self.send_json_data(status=True,data="开始上传")
        self.recv_save_file(target_file_path)

    # 下载文件，断点续传
    def download(self,file_path,seek=0):
        if not self.username:
            self.send_json_data(status=False, error="登录后才能下载")
            return

        # 文件不存在
        target_file_path = os.path.join(self.home_path(),file_path)
        if not os.path.exists(target_file_path):
            self.send_json_data(status=False, error="文件{}不存在".format(file_path))
            return

        self.send_json_data(status=False, error="开始下载{}".format(file_path))

        seek = int(seek)
        total_size = os.stat(target_file_path).st_size
        req.send_file_by_seek(self.conn,total_size-seek,target_file_path,seek)

    def execute(self):
        """
        每次客户端发来请求，触发此方法。
        :return: False，关闭连接；True，继续处理请求
        """

        conn = self.conn

        # 1.获取数据包
        cmd = req.recv_data(conn).decode('utf-8')
        if cmd.upper() == "Q":
            print("客户端退出")
            return False

        """
        login wupeiqi 123
        register wupeiqi 123
        ls   ls xxx（登录成功之后）
        upload v4.py（登录成功之后）
        download v1.txt（登录成功之后）
        """

        method_map = {
            "login": self.login,
            "register": self.register,
            "ls": self.ls,
            "upload": self.upload,
            "download": self.download,
        }

        cmd, *args = re.split(r"\s+", cmd)
        method = method_map[cmd]

        method(*args)

        return True

